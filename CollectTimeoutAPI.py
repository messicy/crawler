import json
from openpyxl import workbook
import time
import hockeyapputils

appid = hockeyapputils.GetAppID()
print(appid)

appversion = hockeyapputils.GetAppVersionID(appid, "1.0.35", "397")
appversion = 19   #temp assign
print(appversion)

page = 0

apidict = dict()

while page < 10000:
    page += 1
    searchurl = 'created_at%3A%5B"2016-05-19T00%3A00"+TO+"2016-05-30T23%3A59"%5D+reason%3A+errmsg%2Bis%2Btimeout%2Btime%2Bis15'
    crashgroupinfostr = hockeyapputils.GetSpecificCrashGroup(appid, appversion, page, searchurl)
    print(crashgroupinfostr)
    time.sleep(1)

    crashgroupinfoobj = json.loads(crashgroupinfostr)
    reasongroup = crashgroupinfoobj["crash_reasons"]
    if len(reasongroup) == 0:
        break

    track = 0
    for reason in reasongroup:
        apistr= reason["reason"]
        apistr = apistr.replace("[Request Error]Request Action is =", "")
        apistr = apistr.replace(" errmsg is timeout time is15", "")
        num = reason["number_of_crashes"]
        if apistr not in apidict:
            apidict[apistr] = num
        else:
            print("duplicate api" + apistr)

# for execl
index = 1
filename = 'timeoutapi.xlsx'
wb = workbook.Workbook()
ws = wb.active
for api in apidict.keys():
    ws['A' + str(index)] = api
    ws['B' + str(index)] = apidict[api]
    index += 1
wb.save(filename)

