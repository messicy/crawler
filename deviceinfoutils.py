import urllib.request
import re
from openpyxl import load_workbook
from urllib.parse import quote

def GenerateDeviceInfoExecl(filename):
    wb = load_workbook(filename)
    ws = wb.active

    index = 1
    urlpattern = re.compile(r'"http:\\\\\/\\\\\/www.devicespecifications.com\\\\\/en\\\\\/model\\\\\/[a-zA-Z0-9]{8}"')
    socpattern = re.compile(r'SoC:( .*?),')
    cpupattern = re.compile(r'CPU:( .*?),')
    gpupattern = re.compile(r'GPU:( .*?),')
    rampattern = re.compile(r'RAM:( .*?),')

    for cell in ws.columns[0]:
        if ws['E' + str(index)].value:
            index = index + 1
            print(index)
            continue

        targetdevice = cell.value
        print(targetdevice)
        if targetdevice:
            url = "http://www.devicespecifications.com/index.php?action=search&language=en&search=" + quote(targetdevice)
            print(url)
            content = urllib.request.urlopen(url).read()
            print(content)
            if len(content) <= 2:
                targetdevicename = str(targetdevice)[:-1]
                content = urllib.request.urlopen(
                    "http://www.devicespecifications.com/index.php?action=search&language=en&search=" + targetdevicename).read()
                print("tt" + str(content))

            if len(content) > 2:
                finalurl = urlpattern.search(str(content))
                if finalurl:
                    finalurlstr = finalurl.group(0)
                    print(finalurlstr)
                    finalurlstr = finalurlstr.replace('\\', '')
                    finalurlstr = finalurlstr.replace('"', '')
                    print(finalurlstr)
                    finalcontent = urllib.request.urlopen(finalurlstr).read()

                    soc = socpattern.search(str(finalcontent))
                    if soc:
                        socstr = soc.group(0)
                        socstr = socstr.replace('SoC: ', '')
                        socstr = socstr.replace(',', '')
                        print(socstr)
                        ws['E' + str(index)] = socstr

                    cpu = cpupattern.search(str(finalcontent))
                    if cpu:
                        cpustr = cpu.group(0)
                        cpustr = cpustr.replace('CPU: ', '')
                        cpustr = cpustr.replace(',', '')
                        print(cpustr)
                        ws['F' + str(index)] = cpustr

                    gpu = gpupattern.search(str(finalcontent))
                    if gpu:
                        gpustr = gpu.group(0)
                        gpustr = gpustr.replace('GPU: ', '')
                        gpustr = gpustr.replace(',', '')
                        print(gpustr)
                        ws['G' + str(index)] = gpustr

                    ram = rampattern.search(str(finalcontent))
                    if ram:
                        ramstr = ram.group(0)
                        ramstr = ramstr.replace('RAM: ', '')
                        ramstr = ramstr.replace(',', '')
                        print(ramstr)
                        ws['H' + str(index)] = ramstr

                    wb.save(filename)

        index = index + 1
        print(index)


