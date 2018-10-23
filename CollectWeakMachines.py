import json
from openpyxl import workbook
import time
import deviceinfoutils
import hockeyapputils

appid = hockeyapputils.GetAppID()
print(appid)

appversion = hockeyapputils.GetAppVersionID(appid, "1.0.35", "397")
appversion = 19   #temp assign
print(appversion)


devicedict = dict()

page = 0
while page < 10000:
    page += 1
    searchurl = 'reason: java\.lang\.Error'
    crashgroupinfostr = hockeyapputils.GetSpecificCrashGroup(appid, appversion, page, searchurl)
    print(crashgroupinfostr)
    time.sleep(1)

    crashgroupinfoobj = json.loads(crashgroupinfostr)
    reasongroup = crashgroupinfoobj["crash_reasons"]
    if len(reasongroup) == 0:
        break

    for reason in reasongroup:
        reasonclass = reason["class"]
        if "libGLES" in reasonclass:
            # print(reasonclass)
            time.sleep(1)

            crashesinfostr = hockeyapputils.GetSpecificCrash(appid, reason["id"])
            # print(crashesinfostr)
            crashesinfoobj = json.loads(crashesinfostr)
            for crash in crashesinfoobj["crashes"]:
                devicekey = crash["model"]
                if devicekey not in devicedict:
                    devicedict[devicekey] = -1
                else:
                    devicedict[devicekey] -= 1

page = 0

while page < 10000:
    page += 1
    searchurl = 'reason: EGL'
    crashesinfostr = hockeyapputils.GetSpecificCrashes(appid, appversion, page, searchurl)
    print(crashesinfostr)
    time.sleep(1)

    crashesinfoobj = json.loads(crashesinfostr)
    reasongroup = crashesinfoobj["crashes"]
    if len(reasongroup) == 0:
        break

    for reason in reasongroup:
        devicekey = reason["model"]
        if devicekey not in devicedict:
            devicedict[devicekey] = 1
        else:
            devicedict[devicekey] += 1

#for execl
index = 1
filename = 'weakmachines.xlsx'
wb = workbook.Workbook()
ws = wb.active
for device in devicedict.keys():
    ws['A' + str(index)] = device
    ws['B' + str(index)] = devicedict[device]
    index += 1
wb.save(filename)

deviceinfoutils.GenerateDeviceInfoExecl(filename)
